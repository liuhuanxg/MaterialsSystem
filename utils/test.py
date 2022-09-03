#!/usr/bin/env python
# -*- coding: utf-8 -*-

import traceback

import xlrd


def parse_excel_data(file_path):
    try:
        wb = xlrd.open_workbook(file_path)
        sh1 = wb.sheet_by_index(0)
        print(u"sheet %s 共 %d 行 %d 列" % (sh1.name, sh1.nrows, sh1.ncols))
        # 打印获取的行列值
        total_rows = sh1.nrows
        total_ncols = sh1.ncols
        if total_rows > 1 and total_ncols >= 4:
            for row_index in range(total_rows):
                row_data = sh1.row_values(row_index)
                materials_name = row_data[0]
                unit = row_data[1]
                specifications = row_data[2]
                number = row_data[3]
                # print(materials_name, unit, specifications, number)
                yield materials_name, unit, specifications, number
    except:
        print("parse_excel_data error:{}".format(traceback.format_exc()))
        return None


from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font


def save_excel():
    wb = Workbook()
    sheet1 = wb.active
    wb.remove(sheet1)
    wb.create_sheet("入库单")
    # Font
    ws = wb["入库单"]
    titles = ["日期", "入库单号", "口罩_n95_个", "手套_n95_个"]
    records_list = {
        "1": {
            "2022-08-30_20220830": {
                "口罩_n95_个": 2,
                "手套_n95_个": 10
            },
            "2022-08-29_20220830": {
                "口罩_n95_个": 2,
                "手套_n95_个": 10
            }

        }
    }
    datas = [titles]
    print(records_list)
    #  开始批量写入数据
    title_tatol = {}
    for _key, data in records_list["1"].items():
        date, app_code = _key.split("_")
        row = []
        row.extend([date, app_code])
        for title in titles[2:]:
            cell_data = data.get(title, "")
            row.append(cell_data)
            if cell_data:
                title_tatol[title] = title_tatol.get(title, 0) + cell_data
        datas.append(row)
    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value='中央库').alignment = Alignment(
        horizontal='center', vertical='center'
    )
    ws["A1"].font = Font(bold=True, size=16)

    for data in datas:
        ws.append(data)

    ws.merge_cells('A{}:B{}'.format(len(datas) + 5, len(datas) + 5))
    ws.cell(row=len(datas) + 5, column=1, value='本期累计').alignment = Alignment(horizontal='center', vertical='center')
    index = 2
    for title in titles[index:]:
        ws.cell(row=len(datas) + 5, column=index + 1, value=title_tatol.get(title, ""))
        index += 1
    wb.save("test3.xlsx")


def test():
    i = 0
    s = ["1", "2", "3", "4", "5"]
    step = 2
    while i < len(s):
        print(s[i:i + step])
        i += step
    print(s[len(s) - i:len(s)])


if __name__ == '__main__':
    # for data in parse_excel_data("./防疫物品.xlsx"):
    #     print(data)
    # save_excel()
    test()
