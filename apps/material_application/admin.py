from center_library.models import CenterOutboundOrder, CenterOutboundOrderDetail
from django.contrib import admin
from django.contrib.admin import DateFieldListFilter
from django.shortcuts import HttpResponse
from django.utils.decorators import method_decorator
from django.utils.html import format_html
from django.views.decorators.csrf import csrf_protect
from home.models import CodeNumber
from local_library.models import LocalOutboundOrder, LocalOutboundOrderDetail, LocalLabraryMaterials
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from MaterialsSystem.settings import status_choices_dict
from .models import *

logger = logging.getLogger("django")


# 审批记录，不允许新增、修改、删除
class ApplicationHistoryInline(admin.TabularInline):
    model = ApplicationHistory
    extra = 0

    def has_add_permission(self, request, obj):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


# 申请详情
class ApplicationDetailInline(admin.TabularInline):
    model = ApplicationDetail
    extra = 0
    fields = ["type_name", "number"]

    def has_change_permission(self, request, obj=None):
        if obj and obj.status == "3":
            return False
        return super(ApplicationDetailInline, self).has_change_permission(request, obj)

    def has_add_permission(self, request, obj):
        if obj and obj.status == "3":
            return False
        return super(ApplicationDetailInline, self).has_add_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.status == "3":
            return False
        return super().has_delete_permission(request, obj)


# 申请时上传的附件
class ApplicationFileInline(admin.TabularInline):
    model = ExApplicationFile
    extra = 0
    fields = ["file"]

    def has_change_permission(self, request, obj=None):
        if obj and obj.status == "3":
            return False
        return super().has_change_permission(request, obj)

    def has_add_permission(self, request, obj):
        if obj and obj.status == "3":
            return False
        return super().has_add_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.status == "3":
            return False
        return super().has_delete_permission(request, obj)


# 地方库研判详情
class LocalAssessmentDetailInline(admin.TabularInline):
    model = LocalAssessmentDetail
    extra = 0
    fields = ["library_name", "number"]
    autocomplete_fields = ["library_name"]

    # 地方库研判，只能选已经审核过的
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "library_name":
            local_labrary_materials = LocalLabraryMaterials.objects.filter(library_name__is_approve=1)
            kwargs["queryset"] = local_labrary_materials
            kwargs['initial'] = local_labrary_materials.first()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def has_change_permission(self, request, obj=None):
        if obj and obj.status == "4":
            return False
        return super().has_change_permission(request, obj)

    def has_add_permission(self, request, obj):
        if obj and obj.status == "4":
            return False
        return super().has_add_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.status == "4":
            return False
        return super().has_delete_permission(request, obj)


# 中央库研判详情
class CenterAssessmentDetailInline(admin.TabularInline):
    model = CenterAssessmentDetail
    extra = 0
    fields = ["library_name", "number"]
    autocomplete_fields = ["type_name"]

    def has_change_permission(self, request, obj=None):
        if obj and obj.status == "4":
            return False
        return super().has_change_permission(request, obj)

    def has_add_permission(self, request, obj):
        if obj and obj.status == "4":
            return False
        return super().has_add_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if obj and obj.status == "4":
            return False
        return super().has_delete_permission(request, obj)


@admin.register(ExWarehousingApplication)
class ExWarehousingApplicationAdmin(admin.ModelAdmin):
    list_display = ["app_code", "title", "applicant", "applicant_user", "add_time",
                    "status_short", "next_node_short"]
    list_filter = [
        "title", "create_user", "applicant_user",
        ("add_date", DateFieldListFilter)
    ]
    date_hierarchy = "add_date"
    readonly_fields = ["create_user"]
    inlines = [ApplicationFileInline, ApplicationDetailInline]
    fields = [("title", "des"), ("applicant", "applicant_user"), ("add_time")]
    db_name = "ExWarehousingApplication"
    change_form_template = "material_application/ex_app_change_form.html"

    def get_inlines(self, request, obj):
        user = request.user
        inlines = self.inlines
        if obj:
            if str(obj.status) and (
                    int(obj.status) == 3 and user.groups.filter(name__contains="主管科室").exists() or int(obj.status) > 3):
                if LocalAssessmentDetailInline not in inlines:
                    inlines.append(LocalAssessmentDetailInline)
                if CenterAssessmentDetailInline not in inlines:
                    inlines.append(CenterAssessmentDetailInline)
            elif not obj.status or int(obj.status) < 3:
                if LocalAssessmentDetailInline in inlines:
                    inlines.remove(LocalAssessmentDetailInline)
                if CenterAssessmentDetailInline in inlines:
                    inlines.remove(CenterAssessmentDetailInline)
            if obj.next_node != "" and ApplicationHistoryInline not in inlines:
                inlines.append(ApplicationHistoryInline)
            logger.info("obj:{}, obj.status:{}, obj.next_node:{}, inlines:{}".format(
                obj, obj.status, obj.next_node, inlines
            ))
        else:
            if ApplicationHistoryInline in inlines:
                inlines.remove(ApplicationHistoryInline)
            if LocalAssessmentDetailInline in inlines:
                inlines.remove(LocalAssessmentDetailInline)
            if CenterAssessmentDetailInline in inlines:
                inlines.remove(CenterAssessmentDetailInline)
        return inlines

    def status_short(self, obj):
        return status_choices_dict.get(obj.status)

    status_short.short_description = u'申请状态'

    def next_node_short(self, obj):
        next_user = User.objects.filter(id=obj.next_node).first()
        if next_user:
            return next_user.username
        return ""

    next_node_short.short_description = u'待审批人'

    def app_status_short(self, obj):
        return format_html(
            '<a href="{}?_id={}">{}</a>'.format("/material_application/do_approval/", obj.id, "点击审批", )
        )

    app_status_short.short_description = u'申请状态'

    def save_formset(self, request, form, formset, change):

        application_id = form.instance.id
        application_status = form.instance.status
        user = request.user
        logger.info("form.instance.status:{}, self_id:{}".format(application_status, application_id))
        if change:
            if application_status == "3":
                # 判断研判的数据结果与真实申请是否相等
                flag = False
                all_applications_details = {}
                application_details = ApplicationDetail.objects.filter(application_id=application_id)
                for application_detail in application_details:
                    type_name_id = application_detail.type_name_id
                    all_applications_details[type_name_id] = all_applications_details.get(type_name_id,
                                                                                          0) + application_detail.number
                logger.info(
                    "application_id:{},all_applications_details:{}".format(application_id, all_applications_details))

                all_assessment_details = {}
                for inline_form in formset.forms:
                    if inline_form.has_changed():
                        if inline_form._meta.model == LocalAssessmentDetail:
                            flag = True
                            type_name_id = inline_form.instance.library_name.type_name_id
                            number = inline_form.instance.number
                            all_assessment_details[type_name_id] = all_assessment_details.get(type_name_id, 0) + number
                        if inline_form._meta.model == CenterAssessmentDetail:
                            flag = True
                            type_name_id = inline_form.instance.library_name.type_name_id
                            number = inline_form.instance.number
                            all_assessment_details[type_name_id] = all_assessment_details.get(type_name_id, 0) + number
                logger.info(
                    "application_id:{},all_assessment_details:{}".format(application_id, all_assessment_details))
                # if flag and all_assessment_details != all_applications_details:
                #     raise ValidationError({"error_dict": "研判数量和申请数量不符。"})
            # 审批过程中校验数据是否变化
            # if user.id != form.instance.create_user_id:
            #     for inline_form in formset.forms:
            #         if inline_form.has_changed() and inline_form._meta.model == ApplicationDetail:
            #             number = inline_form.instance.number
            #             app_detail = ApplicationDetail.objects.filter(id=inline_form.instance.id).first()
            #             if app_detail:
            #                 if app_detail.number > number:
            #                     return ValidationError({"error_dict": "申请数量不能增大。"})
            #             else:
            #                 return ValidationError({"error_dict": "不能新添加物资。"})
            #     logger.info("user.id:{},form.instance.create_user_id:{}".format(user.id, form.instance.create_user_id))
        super(ExWarehousingApplicationAdmin, self).save_formset(request, form, formset, change)

    def save_model(self, request, obj, form, change):
        user = request.user
        application_id = obj.id
        application_user = user.first_name
        action = ""
        logger.info("obj_id:{}, status:{}, next_node:{}, user_id:{}".format(
            obj.id, obj.status, obj.next_node, user.id
        ))

        # 分管领导审批 —— 局长审批 —— 主管科室研判 —— 研判之后生成出库单 —— 中央库管理员出库
        if form.changed_data or obj.next_node == str(user.id) or obj.create_user == user:
            if obj.app_code == "":
                obj.create_user = request.user
                obj.app_code = CodeNumber.get_app_code(self.db_name)
            if obj.create_user == request.user:
                next_user = User.objects.filter(groups__name__icontains="分管领导").first()
                if next_user:
                    next_node = next_user.id
                else:
                    admin = next_user.objects.filter(username__icontains="admin").first()
                    if admin:
                        next_node = admin.id
                    else:
                        next_node = request.user.id
                obj.next_node = next_node
                obj.status = "1"

                action = "发起申请"
            elif obj.next_node == str(request.user.id):
                if obj.status == "1":
                    next_user = User.objects.filter(groups__name__contains="局长").first()
                    if next_user:
                        obj.status = "2"
                        obj.next_node = next_user.id
                        action = "通过"
                elif obj.status == "2":
                    next_user = User.objects.filter(groups__name__contains="主管科室").first()
                    if next_user:
                        obj.status = "3"
                        obj.next_node = next_user.id
                        action = "通过"
                elif obj.status == "3":
                    next_user = User.objects.filter(groups__name__contains="仓库管理员").first()
                    if next_user:
                        obj.status = "4"
                        obj.next_node = next_user.id
                        action = "研判完成"
                # elif obj.status == "5":
                #     obj.status = "6"
                #     user = User.objects.filter(groups__name__contains="仓库管理员").first()
                #     if user:
                #         obj.next_node = user.id
                #         application_user = user.first_name,
                #         action = "通过"
            super(ExWarehousingApplicationAdmin, self).save_model(request, obj, form, change)
            logger.info(application_id, application_user, action)
            ApplicationHistory.objects.create(
                application_id=obj.id,
                application_user=application_user,
                action=action
            )
            # 研判之后生成对应的出库单
            if obj.status == "4":
                # 1.生成中央库出库单
                center_assement_details = CenterAssessmentDetail.objects.filter(application_id=obj.id)
                logger.info("center_assement_details:{}".format(center_assement_details))
                if center_assement_details.exists():
                    # (1) 创建大的出库单子
                    center_outbound_order, err = CenterOutboundOrder.objects.get_or_create(
                        app_code_id=obj.id
                    )
                    center_outbound_order.title = obj.title
                    center_outbound_order.applicant = obj.applicant
                    center_outbound_order.applicant_user = obj.applicant_user

                    center_outbound_order.des = obj.des
                    center_outbound_order.add_time = obj.add_time
                    center_outbound_order.add_date = obj.add_date
                    center_outbound_order.save()
                    # (2) 生成出库小单子
                    for center_assement_detail in center_assement_details:
                        center_outbound_oerder_detail, err = CenterOutboundOrderDetail.objects.get_or_create(
                            app_code_id=center_outbound_order.id,
                            assessment_detail_id=center_assement_detail.id,
                        )
                        center_outbound_oerder_detail.number = center_assement_detail.number
                        center_outbound_oerder_detail.total_price = center_assement_detail.number * center_assement_detail.library_name.unit_price
                        center_outbound_oerder_detail.save()
                        center_outbound_order.total_price = center_outbound_order.total_price + center_outbound_oerder_detail.total_price
                    center_outbound_order.save()

                # 2.生成地方库出库单
                local_assement_details = LocalAssessmentDetail.objects.filter(application_id=obj.id)
                if local_assement_details.exists():
                    total_price = {}
                    for local_assement_detail in local_assement_details:
                        # (1) 创建大的地方库出库单子
                        local_outbound_order, err = LocalOutboundOrder.objects.get_or_create(
                            app_code_id=obj.id,
                            user_id=local_assement_detail.library_name.library_name.supplier_name.user.id,
                        )
                        local_outbound_order.title = obj.title
                        local_outbound_order.applicant = obj.applicant
                        local_outbound_order.applicant_user = obj.applicant_user
                        local_outbound_order.des = obj.des
                        local_outbound_order.add_time = obj.add_time
                        local_outbound_order.add_date = obj.add_date
                        local_outbound_order.save()

                        local_outbound_oerder_detail, err = LocalOutboundOrderDetail.objects.get_or_create(
                            app_code_id=local_outbound_order.id,
                            assessment_detail_id=local_assement_detail.id,
                        )
                        local_outbound_oerder_detail.number = local_assement_detail.number
                        price = local_assement_detail.number * local_assement_detail.library_name.unit_price
                        local_outbound_oerder_detail.total_price = price
                        # 供应商总金额
                        local_outbound_order.total_price = local_outbound_order.total_price + price
                        local_outbound_oerder_detail.save()
                        local_outbound_order.save()

        super(ExWarehousingApplicationAdmin, self).save_model(request, obj, form, change)

    def save_related(self, request, form, formsets, change):
        super(ExWarehousingApplicationAdmin, self).save_related(request, form, formsets, change)

    def response_add(self, request, obj, post_url_continue=None):
        return super(ExWarehousingApplicationAdmin, self).response_add(request, obj, post_url_continue)

    def response_change(self, request, obj):
        return super(ExWarehousingApplicationAdmin, self).response_change(request, obj)

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        self.inlines = [ApplicationFileInline, ApplicationDetailInline]
        extra_context = extra_context if extra_context else {}
        if object_id:
            user = request.user
            obj = self.model.objects.get(id=object_id)
            logger.info("changeform_view self:{}, object_id:{} user_id:{}, next_node:{}, status:{}".format(
                self, object_id, user.id, obj.next_node, obj.status)
            )
            if str(user.id) == obj.next_node:
                extra_context["can_approve"] = True
            else:
                extra_context["can_approve"] = False
        else:
            extra_context["add_application"] = True
        return super(ExWarehousingApplicationAdmin, self).changeform_view(request, object_id, form_url, extra_context)

    def get_inline_formsets(self, request, formsets, inline_instances, obj=None):
        return super(ExWarehousingApplicationAdmin, self).get_inline_formsets(request, formsets, inline_instances, obj)

    def get_formsets_with_inlines(self, request, obj=None):
        for inline in self.get_inline_instances(request, obj):
            # logger.info("get_formsets_with_inlines", inline, inline.get_formset(request, obj))
            yield inline.get_formset(request, obj), inline

    def has_change_permission(self, request, obj=None):
        user = request.user
        if obj:
            if (obj.next_node == str(user.id) or obj.create_user == user) and obj.status and int(obj.status) <= 3:
                return True
            return False
        return super(ExWarehousingApplicationAdmin, self).has_change_permission(request, obj)

    def get_changelist(self, request, **kwargs):
        return super(ExWarehousingApplicationAdmin, self).get_changelist(request, **kwargs)

    def get_changelist_form(self, request, **kwargs):
        return super(ExWarehousingApplicationAdmin, self).get_changelist_form(request, **kwargs)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context if extra_context else {}
        response = super().changelist_view(request, extra_context)
        return response


csrf_protect_m = method_decorator(csrf_protect)


@admin.register(Accounts)
class AccountsaAdmin(admin.ModelAdmin):
    list_filter = [
        "db_type", "supplier_name", "action", "type_name", "applicant",
        ("add_date", DateFieldListFilter)
    ]

    list_display = ["app_code", "db_type", "supplier_name", "entry_name", "type_name",
                    "specifications_des", "unit", "number", "price", "applicant",
                    "unit_price", "add_date", "action", ]
    change_list_template = "material_application/accounts_change_list.html"

    actions = ['download_accounts']
    list_per_page = 50
    date_hierarchy = "add_date"

    def specifications_des(self, obj):
        return obj.specifications[0:5] + "..."

    specifications_des.short_description = u'物料规格'

    def download_accounts(self, request, queryset):
        # TODO 中央库和地方库在一起时要分开导出，或者导出一个压缩包？直接只允许导出一种，多种时就只导出地方库。
        #  开始批量写入数据
        wb = Workbook()

        # 删除自带的sheet
        sheet1 = wb.active
        wb.remove(sheet1)

        # 所有库类型的总计
        db_all_datas = {}
        all_datas = {}
        sheet1_titles = ["日期", "单号", "性质", "单位"]
        total_type_datas = {}
        all_material_sum_count = {}
        col_index = 4
        for record in queryset.values():
            db_type = record["db_type"]
            if db_type not in db_all_datas:
                db_all_datas[db_type] = {}
                all_material_sum_count[db_type] = {}
            all_datas = db_all_datas[db_type]
            material_sum_count = all_material_sum_count[db_type]
            _key = record["type_name"] + "_" + record["specifications"] + "_" + record["unit"]
            _key = _key.replace("/", "")
            date_key = str(record["add_date"]) + "_" + str(record["app_code"])
            action = record["action"]
            number = record["number"]
            if action not in all_datas:
                all_datas[action] = {}
            if date_key not in all_datas[action]:
                all_datas[action][date_key] = {}
            all_datas[action][date_key][_key] = all_datas[action][date_key].get(_key, 0) + number
            if _key not in sheet1_titles:
                sheet1_titles.append(_key)

            if _key not in total_type_datas:
                total_type_datas[_key] = {}
            if date_key not in total_type_datas[_key]:
                total_type_datas[_key][date_key] = {}
            if action not in total_type_datas[_key][date_key]:
                total_type_datas[_key][date_key][action] = record

            if _key not in material_sum_count:
                material_sum_count[_key] = {}
            material_sum_count[_key][action] = material_sum_count[_key].get(action, 0) + number

        if "1" in db_all_datas:
            all_datas = db_all_datas["1"]
            material_sum_count = all_material_sum_count["1"]
        else:
            all_datas = db_all_datas["2"]
            material_sum_count = all_material_sum_count["2"]
        # 组装入库单sheet，前三个固定，后边的一些按照typename生成
        insert = "物资入库"
        out = "物资出库"
        insert_and_out = "物资结存汇总表"
        all_sheets = [insert, out, insert_and_out]
        wb.create_sheet(insert)
        wb.create_sheet(out)
        wb.create_sheet(insert_and_out)
        insert_sheet = wb[insert]
        out_sheet = wb[out]
        out_and_insert = wb[insert_and_out]

        insert_sheet_data = [sheet1_titles]
        out_sheet_data = [sheet1_titles]
        # 创建总的入库和出库台账
        for action, items in all_datas.items():
            for _key, data in items.items():
                date, app_code = _key.split("_")
                row = [date, app_code, "", ""]
                for title in sheet1_titles[col_index:]:
                    row.append(data.get(title, ""))
                if action == "1":
                    insert_sheet_data.append(row)
                else:
                    out_sheet_data.append(row)

        # 修改表头，合并若干列
        insert_sheet.merge_cells("A1:K1")
        insert_sheet.cell(row=1, column=1, value='入库记录').alignment = Alignment(
            horizontal='center', vertical='center'
        )
        insert_sheet["A1"].font = Font(bold=True, size=16)

        for insert_data in insert_sheet_data:
            insert_sheet.append(insert_data)
        insert_sheet.merge_cells('A{}:D{}'.format(len(insert_sheet_data) + 5, len(insert_sheet_data) + 5))
        insert_sheet.cell(row=len(insert_sheet_data) + 5, column=1, value='本期累计').alignment = Alignment(
            horizontal='center', vertical='center'
        )

        out_sheet.merge_cells("A1:K1")
        out_sheet.cell(row=1, column=1, value='出库记录').alignment = Alignment(
            horizontal='center', vertical='center'
        )
        out_sheet["A1"].font = Font(bold=True, size=16)
        for out_data in out_sheet_data:
            out_sheet.append(out_data)
        out_sheet.merge_cells('A{}:D{}'.format(len(out_sheet_data) + 5, len(out_sheet_data) + 5))
        out_sheet.cell(row=len(out_sheet_data) + 5, column=1, value='出库记录').alignment = Alignment(
            horizontal='center', vertical='center'
        )

        # 补充入库和出库总计数量
        for index in range(col_index, len(sheet1_titles[col_index:])):
            title = sheet1_titles[index]
            insert_value = material_sum_count.get(title, {}).get("1", "")
            insert_sheet.cell(row=len(insert_sheet_data) + 5, column=index + 1,
                              value=insert_value)
            out_value = material_sum_count.get(title, {}).get("2", "")
            logger.info(
                "title:{},insert_value:{},out_value:{}, row:{}, column:{}".format(
                    title, insert_value, out_value, len(out_sheet_data) + 5, index + 1)
            )
            out_sheet.cell(row=len(out_sheet_data) + 5, column=index + 1,
                           value=out_value)

        # 创建物资汇总sheet
        total_titles = ["序号", "物资名称", "规格", "单位", "入库数量", "出库数量", "结存数量"]
        insert_and_out_datas = []
        insert_and_out_datas.append(total_titles)
        logger.info("insert_and_out_datas:{}".format(insert_and_out_datas))
        material_id = 1
        for materials, counts in material_sum_count.items():
            # logger.info("materials:{}".format(materials))
            type_name, specifications, unit = materials.split("_")
            insert_and_out_datas.append(
                [material_id, type_name, specifications, unit, counts.get("1", ""), counts.get("2", ""),
                 counts.get("1", 0) - counts.get("2", 0)]
            )
            material_id += 1
        out_and_insert.merge_cells("A1:G1")
        out_and_insert.cell(row=1, column=1, value='汇总记录').alignment = Alignment(
            horizontal='center', vertical='center'
        )
        out_and_insert["A1"].font = Font(bold=True, size=16)

        for insert_and_out_data in insert_and_out_datas:
            out_and_insert.append(insert_and_out_data)

        # 创建各个类型的分类台账
        for type_name, items in total_type_datas.items():
            all_sheets.append(type_name)
            wb.create_sheet(type_name)
            ws = wb[type_name]
            ws.merge_cells("A1:S1")
            ws.cell(row=1, column=1, value=type_name).alignment = Alignment(
                horizontal='center', vertical='center'
            )
            ws["A1"].font = Font(bold=True, size=16)
            titles = ["日期", "单号", "领用单位", "规格型号", "采购数量", "单价(元)", "金额(元)", "领用数量", "结存"]
            type_names = type_name.split("_")
            ws.append(titles)
            for date_key, data in items.items():
                date, app_code = date_key.split("_")
                row = [date, app_code]
                for action, record in data.items():
                    if action == "2":
                        row.extend(
                            [record["applicant"], type_names[1], "", record["unit_price"], "", record["number"], ""])
                    else:
                        row.extend(["", type_names[1], record["number"], record["unit_price"], record["price"], "", ""])
                ws.append(row)
            ws.append([
                "", "", "本期累计", "",
                material_sum_count.get(type_name, {}).get("1", ""),
                "",
                "",
                material_sum_count.get(type_name, {}).get("2", ""),
                material_sum_count.get(type_name, {}).get("1", 0) - material_sum_count.get(type_name, {}).get("2", 0),
            ])
        # 自动调整列宽
        lks = []
        for sheet in all_sheets:
            ws = wb[sheet]
            for i in range(1, ws.max_column + 1):  # 每列循环
                lk = 1  # 定义初始列宽，并在每个行循环完成后重置
                cell = ws.cell(row=2, column=i)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', text_rotation=0, wrap_text=True
                )
                sz = cell.value  # 每个单元格内容
                if isinstance(sz, str):  # 中文占用多个字节，需要分开处理
                    lk1 = len(sz.encode('gbk'))  # gbk解码一个中文两字节，utf-8一个中文三字节，gbk合适
                else:
                    lk1 = len(str(sz))
                if lk < lk1:
                    lk = lk1  # 借助每行循环将最大值存入lk中
                lks.append(lk)  # 将每列最大宽度加入列表。（犯了一个错，用lks = lks.append(lk)报错，append会修改列表变量，返回值none，而none不能继续用append方法）

            # 第二步：设置列宽
            for i in range(1, ws.max_column + 1):
                # 将数字转化为列名,26个字母以内也可以用[chr(i).upper() for i in range(97, 123)]，不用导入模块
                k = get_column_letter(i)
                # 设置列宽，一般加两个字节宽度，可以根据实际情况灵活调整
                ws.column_dimensions[k].width = lks[i - 1] + 2 if lks[i - 1] < 10 else lks[i - 1] / 2
                ws.column_dimensions[k].height = 108
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="台账记录.xlsx"'
        wb.save(response)
        return response

    download_accounts.short_description = u'导出台账'

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context if extra_context else {}
        response = super().changelist_view(request, extra_context)
        # logger.info(hasattr(response, "context_data"))
        if hasattr(response, "context_data") and "cl" in response.context_data:
            cl = response.context_data["cl"]
            filtered_query_set = cl.get_queryset(request)
            put_number_sum = 0
            push_number_sum = 0
            for i in filtered_query_set:
                if i.action == "1":
                    put_number_sum += i.number
                else:
                    push_number_sum += i.number
            response.context_data["put_number_sum"] = put_number_sum
            response.context_data["push_number_sum"] = push_number_sum
        return response

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False
